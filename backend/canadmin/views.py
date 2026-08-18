from django.db.models import Prefetch, Sum
from django.db.models.functions import Round
from django.shortcuts import get_object_or_404, render
from django.http import HttpResponse
from .models import AptitudeAnswerModel, CodingQuestionSubmissionModel, CodingSubmissionModel, InstituteModel, DriveModel, RoundCandidateDecisionModel, RoundModel, AptitudeQuestionModel, CodingQuestionModel, RoundAttemptModel, CodingTestCaseModel
from .serializers import InstituteSerializer, DriveSerializer, RoundSerializer, CodingQuestionSerializer, DriveDetailsSerializer, UploadAptitudeQuestionSerializer, AptitudeQuestionSerializer, CodingTestCaseSerializer
from .services import AutoStatusService
from candidate.models import CandidateProfile, Education, Experience, Project, Skill, Certificate, Language
from candidate.serializers import ProjectSerializer, EducationSerializer, ExperienceSerializer, SkillSerializer, CertificateSerializer, LanguageSerializer, CandidateProfileSerializer
from authentication.models import UserTable
from authentication.serializers import UserSerializers
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from openpyxl import load_workbook
from django.db import transaction
from django.utils import timezone
import logging

logger = logging.getLogger(__name__)

# Create your views here.
def test(request):
    return HttpResponse("<h1>This Is Admin Test</h1>")

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def GetUsers(request):
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    users = UserTable.objects.filter(role="candidate").order_by("-date_joined")
    serializer = UserSerializers(users, many=True, context={'request': request})
    return Response(
        {
            "success": True,
            "message": "Users fetched successfully!",
            "count": users.count(),
            "data": serializer.data,
        },
        status=status.HTTP_200_OK,
    )

@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def UpdateUserStatus(request, user_id):
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    try:
        user = UserTable.objects.get(id=user_id, role="candidate")
    except UserTable.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "User Not Found"
            },
            status= status.HTTP_404_NOT_FOUND
        )
    user.is_active = not user.is_active
    user.save()
    return Response(
        {
            "success": True,
            "message": f"User {'activated' if user.is_active else 'deactivated'} successfully!!",
            "data": {
                "id": user.id,
                "is_active": user.is_active
            }
        },
        status= status.HTTP_200_OK
    )

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def GetUserDetails(request, user_id):
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    try:
        user = UserTable.objects.select_related("profile").get(id=user_id, role="candidate")
    except UserTable.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "User Not Found"
            },
            status= status.HTTP_404_NOT_FOUND
        )
    try:
        profile = user.profile
    except CandidateProfile.DoesNotExist:
        profile = None
    response = {
        "user": UserSerializers(user).data,
        "profile": CandidateProfileSerializer(profile).data if profile else None,
        "educations": EducationSerializer(Education.objects.filter(user=user), many= True).data,
        "experiences": ExperienceSerializer(Experience.objects.filter(user=user), many= True).data,
        "skills": SkillSerializer(Skill.objects.filter(user=user), many= True).data,
        "projects": ProjectSerializer(Project.objects.filter(user=user), many= True).data,
        "languages": LanguageSerializer(Language.objects.filter(user=user), many= True).data,
        "certificates": CertificateSerializer(Certificate.objects.filter(user=user), many= True).data,
    }
    return Response(
        {
            "success": True,
            "message": "User Profile Fetched Successfully!!!",
            "data": response
        },
        status= status.HTTP_200_OK
    )

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def AddUpdateInstitute(request):
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )    
    institute_id = request.data.get("id")
    if institute_id:
        try:
            institute = InstituteModel.objects.get(id=institute_id)
        except InstituteModel.DoesNotExist:
            return Response(
                {
                    "success": False,
                    "message": "Institute not found."
                },
                status=status.HTTP_404_NOT_FOUND
            )
        serializer = InstituteSerializer(institute, data=request.data, partial=True)
        if serializer.is_valid():
            institute = serializer.save()
            return Response(
                {
                    "success": True,
                    "message": "Institute Updated Successfully!!",
                    "data": InstituteSerializer(institute).data
                },
                status=status.HTTP_201_CREATED
            )
        return Response(
            {
                "success": False,
                "message": serializer.errors
            },
            status=status.HTTP_400_BAD_REQUEST
        )
    serializer = InstituteSerializer(data=request.data)
    if serializer.is_valid():
        institute = serializer.save()
        return Response(
            {
                "success": True,
                "message": "Institute Added Successfully!!",
                "data": InstituteSerializer(institute).data
            },
            status=status.HTTP_201_CREATED
        )
    return Response(
        {
            "success": False,
            "message": serializer.errors
        },
        status=status.HTTP_400_BAD_REQUEST
    )

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def GetInstitute(request):
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    institute = InstituteModel.objects.all().order_by("created_at")
    serializer = InstituteSerializer(institute, many=True, context={'request': request})
    return Response(
        {
            "success": True,
            "message": "Institute fetched successfully!",
            "count": institute.count(),
            "data": serializer.data,
        },
        status=status.HTTP_200_OK,
    )

@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def UpdateInstituteStatus(request, institute_id):
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    try:
        institute = InstituteModel.objects.get(id=institute_id)
    except InstituteModel.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "Institute Not Found"
            },
            status=status.HTTP_404_NOT_FOUND
        )

    institute.is_active = not institute.is_active
    institute.save()
    return Response(
        {
            "success": True,
            "message": f"Institute {'activated' if institute.is_active else 'deactivated'} successfully!!",
            "data": {
                "id": institute.id,
                "is_active": institute.is_active
            }
        },
        status=status.HTTP_200_OK
    )

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def AddUpdateDrive(request):
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    drive_id = request.data.get("id")
    
    if drive_id:
        try:
            drive = DriveModel.objects.get(id=drive_id)
        except DriveModel.DoesNotExist:
            return Response(
                {
                    "success": False,
                    "message": "Drive not found."
                },
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Prevent status change for completed or cancelled drives
        if drive.status in ["completed", "cancelled"]:
            return Response(
                {
                    "success": False,
                    "message": f"Cannot modify a {drive.status} drive."
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Remove status from data - admin cannot change status directly
        data = request.data.copy()
        if "status" in data:
            data.pop("status")
        
        # Keep the original status
        data["status"] = drive.status
        
        serializer = DriveSerializer(drive, data=data, partial=True)
        if serializer.is_valid():
            drive = serializer.save()
            return Response(
                {
                    "success": True,
                    "message": "Drive Updated Successfully!!",
                    "data": DriveSerializer(drive).data
                },
                status=status.HTTP_200_OK
            )
        return Response(
            {
                "success": False,
                "message": serializer.errors
            },
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # New drive - always starts as draft
    data = request.data.copy()
    data["status"] = "draft"
    
    serializer = DriveSerializer(data=data)
    if serializer.is_valid():
        drive = serializer.save()
        return Response(
            {
                "success": True,
                "message": "Drive Added Successfully!!",
                "data": DriveSerializer(drive).data
            },
            status=status.HTTP_201_CREATED
        )
    return Response(
        {
            "success": False,
            "message": serializer.errors
        },
        status=status.HTTP_400_BAD_REQUEST
    )

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def GetDrive(request):
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    
    # Auto-update statuses before returning data
    AutoStatusService.update_all()
    
    drive = DriveModel.objects.all().order_by("created_at")
    serializer = DriveSerializer(drive, many=True, context={'request': request})
    return Response(
        {
            "success": True,
            "message": "Drive fetched successfully!",
            "count": drive.count(),
            "data": serializer.data,
        },
        status=status.HTTP_200_OK,
    )

@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def UpdateDriveStatus(request, drive_id):
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    try:
        drive = DriveModel.objects.get(id=drive_id)
    except DriveModel.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "Drive not found."
            },
            status=status.HTTP_404_NOT_FOUND
        )
    new_status = request.data.get("status")
    if not new_status:
        return Response(
            {
                "success": False,
                "message": "Status is required."
            },
            status=status.HTTP_400_BAD_REQUEST
        )
    # Define allowed transitions - No reverse transitions allowed
    ALLOWED_TRANSITIONS = {
        'draft': ['published', 'cancelled'],
        'published': ['in_progress', 'cancelled'],
        'in_progress': ['completed', 'cancelled'],
        'completed': [],
        'cancelled': []
    }
    # Check if transition is allowed
    if new_status not in ALLOWED_TRANSITIONS.get(drive.status, []):
        return Response(
            {
                "success": False,
                "message": f'Cannot change status from "{drive.status}" to "{new_status}"'
            },
            status=status.HTTP_400_BAD_REQUEST
        )
    drive.status = new_status
    drive.save(update_fields=["status", "updated_at"])
    return Response(
        {
            "success": True,
            "message": f"Drive status updated to {new_status.replace('_', ' ').title()} successfully!!",
            "data": {
                "id": drive.id,
                "status": drive.status
            }
        },
        status=status.HTTP_200_OK
    )

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def AddUpdateRound(request):
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    round_id = request.data.get("id")
    
    if round_id:
        try:
            round_obj = RoundModel.objects.get(id=round_id)
        except RoundModel.DoesNotExist:
            return Response(
                {
                    "success": False,
                    "message": "Round Not Found!!!"
                },
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Prevent status change for completed or cancelled rounds
        if round_obj.status in ["completed", "cancelled"]:
            return Response(
                {
                    "success": False,
                    "message": f"Cannot modify a {round_obj.status} round."
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Remove status from data - admin cannot change status directly
        data = request.data.copy()
        if "status" in data:
            data.pop("status")
        
        # Keep the original status
        data["status"] = round_obj.status
        
        serializer = RoundSerializer(round_obj, data=data, partial=True)
        if serializer.is_valid():
            round_obj = serializer.save()
            return Response(
                {
                    "success": True,
                    "message": "Round Updated Successfully!!!",
                    "data": RoundSerializer(round_obj).data
                },
                status=status.HTTP_200_OK
            )
        return Response(
            {
                "success": False,
                "message": serializer.errors
            },
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # New round - always starts as pending
    data = request.data.copy()
    data["status"] = "pending"
    
    serializer = RoundSerializer(data=data)
    if serializer.is_valid():
        round_obj = serializer.save()
        return Response(
            {
                "success": True,
                "message": "Round added successfully.",
                "data": RoundSerializer(round_obj).data
            },
            status=status.HTTP_201_CREATED
        )
    return Response(
        {
            "success": False,
            "message": serializer.errors
        },
        status=status.HTTP_400_BAD_REQUEST
    )

@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def UpdateRoundStatus(request, round_id):
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    try:
        round_obj = RoundModel.objects.get(id=round_id)
    except RoundModel.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "Round not found."
            },
            status=status.HTTP_404_NOT_FOUND
        )
    new_status = request.data.get("status")
    if not new_status:
        return Response(
            {
                "success": False,
                "message": "Status is required."
            },
            status=status.HTTP_400_BAD_REQUEST
        )
    # Define allowed transitions - No reverse transitions allowed
    ALLOWED_TRANSITIONS = {
        'pending': ['active', 'cancelled'],
        'active': ['completed', 'cancelled'],
        'completed': [],
        'cancelled': []
    }
    # Check if transition is allowed
    if new_status not in ALLOWED_TRANSITIONS.get(round_obj.status, []):
        return Response(
            {
                "success": False,
                "message": f'Cannot change round status from "{round_obj.status}" to "{new_status}"'
            },
            status=status.HTTP_400_BAD_REQUEST
        )
    round_obj.status = new_status
    round_obj.save(update_fields=["status", "updated_at"])
    return Response(
        {
            "success": True,
            "message": f"Round status updated to {new_status.replace('_', ' ').title()} successfully!!",
            "data": {
                "id": round_obj.id,
                "status": round_obj.status
            }
        },
        status=status.HTTP_200_OK
    )

@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def DeleteRound(request, round_id):
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    try:
        round_obj = RoundModel.objects.get(id=round_id)
    except RoundModel.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "Round not found."
            },
            status=status.HTTP_404_NOT_FOUND
        )
    round_id = round_obj.id
    round_type = round_obj.round_type
    round_obj.delete()
    return Response(
        {
            "success": True,
            "message": (
                f"{round_type.replace('_', ' ').title()} "
                "round deleted successfully!!"
            ),
            "data": {
                "id": round_id
            }
        },
        status=status.HTTP_200_OK
    )

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def GetDriveDetails(request, drive_id):
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    
    # Auto-update the specific drive
    try:
        drive = DriveModel.objects.select_related("institute").prefetch_related("rounds").get(id=drive_id)
        AutoStatusService.update_drive_status(drive)
        # Also update all rounds
        for round_obj in drive.rounds.all():
            AutoStatusService.update_round_status(round_obj)
    except DriveModel.DoesNotExist:
        pass
    
    try:
        drive = DriveModel.objects.select_related("institute").prefetch_related("rounds").get(id=drive_id)
    except DriveModel.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "Drive not found."
            },
            status=status.HTTP_404_NOT_FOUND
        )
    
    try:
        serializer = DriveDetailsSerializer(drive)
        return Response(
            {
                "success": True,
                "data": serializer.data
            },
            status=status.HTTP_200_OK
        )
    except Exception as e:
        logger.error(f"Error serializing drive {drive_id}: {str(e)}")
        return Response(
            {
                "success": False,
                "message": f"Error serializing drive data: {str(e)}"
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def UploadAptitudeQuestion(request):
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    serializer = UploadAptitudeQuestionSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(
            {
                "success": False,
                "message": serializer.errors
            },
            status=status.HTTP_400_BAD_REQUEST
        )
    round_id = serializer.validated_data["round_id"]
    excel_file = serializer.validated_data["file"]
    try:
        round_obj = RoundModel.objects.get(id=round_id)
    except RoundModel.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "Round not found."
            },
            status=status.HTTP_404_NOT_FOUND
        )
    if round_obj.round_type != "aptitude":
        return Response(
            {
                "success": False,
                "message": "Questions can only be uploaded to an aptitude round."
            },
            status=status.HTTP_400_BAD_REQUEST
        )
    workbook = load_workbook(excel_file)
    sheet = workbook.active
    expected_headers = ["Question", "Option 1", "Option 2", "Option 3", "Option 4", "Correct Option", "Marks"]
    headers = [cell.value for cell in sheet[1]]
    if headers != expected_headers:
        return Response(
            {
                "success": False,
                "message": "Invalid Excel format.",
                "expected_headers": expected_headers
            },
            status=status.HTTP_400_BAD_REQUEST
        )
    questions = []
    try:
        with transaction.atomic():
            for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                question = row[0]
                option1 = row[1]
                option2 = row[2]
                option3 = row[3]
                option4 = row[4]
                correct_option = row[5]
                marks = row[6]
                print(f"Row {row_number}: {row}")
                if not question:
                    raise Exception(f"Row {row_number}: Question is empty.")
                options = [option1, option2, option3, option4]
                if any(option is None or (isinstance(option, str) and option.strip() == "") for option in options):
                    raise Exception(f"Row {row_number}: All four options are required.")
                if correct_option not in ["option_1", "option_2", "option_3", "option_4"]:
                    raise Exception(f"Row {row_number}: Invalid Correct Option.")
                if marks is None:
                    marks = 1
                questions.append(AptitudeQuestionModel(round=round_obj, question=question, option_1=option1, option_2=option2, option_3=option3, option_4=option4, correct_option=correct_option, marks=int(marks)))
            AptitudeQuestionModel.objects.bulk_create(questions)
    except Exception as e:
        return Response(
            {
                "success": False,
                "message": str(e)
            },
            status=status.HTTP_400_BAD_REQUEST
        )
    return Response(
        {
            "success": True,
            "message": f"{len(questions)} Questions Uploaded Successfully."
        },
        status=status.HTTP_201_CREATED
    )

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def GetAptitudeQuestions(request, round_id):
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    try:
        round_obj = RoundModel.objects.get(id=round_id)
    except RoundModel.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "Round not found."
            },
            status=status.HTTP_404_NOT_FOUND
        )
    if round_obj.round_type != "aptitude":
        return Response(
            {
                "success": False,
                "message": "This is not an aptitude round."
            },
            status=status.HTTP_400_BAD_REQUEST
        )
    questions = AptitudeQuestionModel.objects.filter(round=round_obj).order_by("id")
    serializer = AptitudeQuestionSerializer(questions, many=True)
    return Response(
        {
            "success": True,
            "message": "Questions fetched successfully.",
            "count": questions.count(),
            "data": serializer.data
        },
        status=status.HTTP_200_OK
    )

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def AddUpdateAptitudeQuestion(request):
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    aptitude_question_id = request.data.get("id")
    if aptitude_question_id:
        try:
            aptitude_question = AptitudeQuestionModel.objects.get(id = aptitude_question_id)
        except AptitudeQuestionModel.DoesNotExist:
            return Response(
                {
                    "success": False,
                    "message": "Aptitude Question Not Found!!!"
                },
                status= status.HTTP_404_NOT_FOUND
            )
        serializer = AptitudeQuestionSerializer(aptitude_question, data= request.data, partial=True)
        if serializer.is_valid():
            aptitude_question = serializer.save()
            return Response(
                {
                    "success": True,
                    "message": "Aptitude Question Updated Successfully!!!",
                    "data": AptitudeQuestionSerializer(aptitude_question).data
                },
                status= status.HTTP_200_OK
            )
        return Response(
            {
                "success": False,
                "message": serializer.errors
            },
            status= status.HTTP_400_BAD_REQUEST
        )
    serializer = AptitudeQuestionSerializer(data=request.data)
    if serializer.is_valid():
        aptitude_question = serializer.save()
        return Response(
            {
                "success": True,
                "message": "Aptitude Question added successfully.",
                "data": AptitudeQuestionSerializer(aptitude_question).data
            },
            status=status.HTTP_201_CREATED
        )
    return Response(
        {
            "success": False,
            "message": serializer.errors
        },
        status=status.HTTP_400_BAD_REQUEST
    )

@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def DeleteAptitudeQuestion(request, aptitude_question_id):
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    try:
        aptitude_question_obj = AptitudeQuestionModel.objects.get(id=aptitude_question_id)
    except AptitudeQuestionModel.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "Aptitude Question Not Found."
            },
            status=status.HTTP_404_NOT_FOUND
        )
    aptitude_question_id = aptitude_question_obj.id
    aptitude_question_obj.delete()
    return Response(
        {
            "success": True,
            "message": "Aptitude Question Deleted Successfully!!",
            "data": {
                "id": aptitude_question_id
            }
        },
        status=status.HTTP_200_OK
    )

# canadmin/views.py - Update ListAptitudeResults

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def ListAptitudeResults(request, round_id):
    """
    Get aptitude results with question-wise details for a specific attempt
    """
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Only admin can view aptitude results."
            },
            status=status.HTTP_403_FORBIDDEN
        )

    try:
        round_obj = RoundModel.objects.get(
            id=round_id,
            round_type="aptitude"
        )
    except RoundModel.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "Aptitude Round Not Found."
            },
            status=status.HTTP_404_NOT_FOUND
        )

    # Get attempt_id from query params for detailed view
    attempt_id = request.query_params.get('attempt_id')
    
    if attempt_id:
        # Return detailed results for a specific attempt
        try:
            attempt = RoundAttemptModel.objects.select_related(
                'candidate', 'round'
            ).get(
                id=attempt_id,
                round=round_obj
            )
            
            # Get all answers for this attempt
            answers = AptitudeAnswerModel.objects.filter(
                attempt=attempt
            ).select_related('question')
            
            # Get questions for this round
            questions = AptitudeQuestionModel.objects.filter(
                round=round_obj
            ).order_by('id')
            
            # Build question-wise results
            question_results = []
            for question in questions:
                answer = answers.filter(question=question).first()
                question_results.append({
                    'question_id': question.id,
                    'question': question.question,
                    'options': [
                        question.option_1,
                        question.option_2,
                        question.option_3,
                        question.option_4
                    ],
                    'correct_option': question.correct_option,
                    'selected_option': answer.selected_option if answer else None,
                    'is_correct': answer.is_correct if answer else False,
                    'marks': question.marks,
                    'marks_obtained': answer.marks_obtained if answer else 0
                })
            
            # Get candidate decision
            candidate_decision = None
            try:
                decision_obj = RoundCandidateDecisionModel.objects.get(attempt=attempt)
                candidate_decision = {
                    'decision': decision_obj.decision,
                    'score': decision_obj.score,
                    'total_marks': decision_obj.total_marks,
                    'percentage': float(decision_obj.percentage) if decision_obj.percentage else 0
                }
            except RoundCandidateDecisionModel.DoesNotExist:
                pass
            
            # Calculate stats
            total_questions = len(question_results)
            correct_answers = sum(1 for q in question_results if q['is_correct'])
            wrong_answers = total_questions - correct_answers - sum(1 for q in question_results if q['selected_option'] is None)
            unattempted = sum(1 for q in question_results if q['selected_option'] is None)
            score = attempt.score or 0
            total_marks = attempt.total_marks or 0
            
            return Response(
                {
                    "success": True,
                    "message": "Aptitude results fetched successfully.",
                    "data": {
                        "attempt_id": attempt.id,
                        "candidate_name": attempt.candidate.name or attempt.candidate.email,
                        "candidate_email": attempt.candidate.email,
                        "score": score,
                        "total_marks": total_marks,
                        "percentage": round((score / total_marks * 100), 2) if total_marks > 0 else 0,
                        "correct_answers": correct_answers,
                        "wrong_answers": wrong_answers,
                        "unattempted": unattempted,
                        "status": attempt.status,
                        "submitted_at": attempt.submitted_at,
                        "candidate_decision": candidate_decision,
                        "questions": question_results
                    }
                },
                status=status.HTTP_200_OK
            )
            
        except RoundAttemptModel.DoesNotExist:
            return Response(
                {
                    "success": False,
                    "message": "Attempt not found."
                },
                status=status.HTTP_404_NOT_FOUND
            )
    
    # Original: Return list of all attempts
    attempts = (
        RoundAttemptModel.objects
        .filter(round=round_obj)
        .select_related("candidate", "round")
        .order_by("-submitted_at", "-started_at")
    )

    attempt_ids = [attempt.id for attempt in attempts]
    decisions = RoundCandidateDecisionModel.objects.filter(
        attempt_id__in=attempt_ids
    ).select_related('attempt')
    decision_map = {decision.attempt_id: decision for decision in decisions}

    results = []
    for attempt in attempts:
        score = attempt.score or 0
        total_marks = attempt.total_marks or 0
        percentage = 0
        if total_marks > 0:
            percentage = round((score / total_marks) * 100, 2)

        candidate_name = getattr(attempt.candidate, "name", None)
        if not candidate_name:
            candidate_name = attempt.candidate.email

        candidate_decision = None
        decision_obj = decision_map.get(attempt.id)
        if decision_obj:
            candidate_decision = {
                "decision": decision_obj.decision,
                "score": decision_obj.score,
                "total_marks": decision_obj.total_marks,
                "percentage": float(decision_obj.percentage) if decision_obj.percentage else 0
            }

        results.append({
            "attempt_id": attempt.id,
            "candidate_id": attempt.candidate.id,
            "candidate_name": candidate_name,
            "candidate_email": attempt.candidate.email,
            "status": attempt.status,
            "score": score,
            "total_marks": total_marks,
            "percentage": percentage,
            "started_at": attempt.started_at,
            "submitted_at": attempt.submitted_at,
            "candidate_decision": candidate_decision,
        })

    total_candidates = len(results)
    passed_count = 0
    failed_count = 0
    pending_count = 0

    for result in results:
        decision = result.get("candidate_decision")
        if decision:
            if decision["decision"] == "shortlisted":
                passed_count += 1
            elif decision["decision"] == "rejected":
                failed_count += 1
            else:
                pending_count += 1
        else:
            pending_count += 1

    return Response(
        {
            "success": True,
            "message": "Aptitude Results Fetched Successfully.",
            "data": {
                "round_id": round_obj.id,
                "round_type": round_obj.round_type,
                "round_order": round_obj.round_order,
                "total_candidates": total_candidates,
                "results": results,
                "summary": {
                    "passed": passed_count,
                    "failed": failed_count,
                    "pending": pending_count
                }
            }
        },
        status=status.HTTP_200_OK
    )

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def AutoUpdateDriveStatuses(request):
    """Auto-update drive statuses based on time and round activity"""
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    
    try:
        now = timezone.now()
        updated_drive_ids = []
        status_changes = []
        
        # 1. AUTO-PUBLISH: Draft → Published when drive date/time arrives
        draft_drives = DriveModel.objects.filter(
            status='draft',
            drive_date_time__lte=now
        )
        for drive in draft_drives:
            drive.status = 'published'
            drive.save(update_fields=["status", "updated_at"])
            updated_drive_ids.append(drive.id)
            status_changes.append(f"Drive {drive.id} auto-published (Draft → Published)")
            logger.info(f"Drive {drive.id} auto-published to published")
        
        # 2. Published → In Progress: When at least one round is active
        published_drives = DriveModel.objects.filter(status='published')
        for drive in published_drives:
            # Check if any round is active
            active_rounds = drive.rounds.filter(status='active')
            if active_rounds.exists():
                drive.status = 'in_progress'
                drive.save(update_fields=["status", "updated_at"])
                updated_drive_ids.append(drive.id)
                status_changes.append(f"Drive {drive.id} auto-updated to in_progress (Round active)")
                logger.info(f"Drive {drive.id} auto-updated to in_progress")
        
        # 3. In Progress → Completed: When all rounds are completed
        in_progress_drives = DriveModel.objects.filter(status='in_progress')
        for drive in in_progress_drives:
            rounds = drive.rounds.all()
            if rounds and all(round.status == 'completed' for round in rounds):
                drive.status = 'completed'
                drive.save(update_fields=["status", "updated_at"])
                updated_drive_ids.append(drive.id)
                status_changes.append(f"Drive {drive.id} auto-updated to completed")
                logger.info(f"Drive {drive.id} auto-updated to completed")
        
        return Response({
            'success': True,
            'updated_drive_ids': updated_drive_ids,
            'status_changes': status_changes,
            'message': f'Updated {len(updated_drive_ids)} drives'
        })
    except Exception as e:
        logger.error(f"Error in auto_update_drive_statuses: {str(e)}")
        return Response({
            'success': False,
            'message': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# canadmin/views.py - Updated AutoUpdateRoundStatuses

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def AutoUpdateRoundStatuses(request):
    """Auto-update round statuses based on time"""
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    
    try:
        now = timezone.now()
        updated_round_ids = []
        drive_id = request.data.get('drive_id')
        
        rounds_query = RoundModel.objects.all()
        if drive_id:
            rounds_query = rounds_query.filter(drive_id=drive_id)
        
        # 1. Update rounds from 'pending' to 'active' when start time arrives
        pending_rounds = rounds_query.filter(
            status='pending',
            round_start_datetime__lte=now
        )
        for round_obj in pending_rounds:
            round_obj.status = 'active'
            round_obj.save(update_fields=["status", "updated_at"])
            updated_round_ids.append(round_obj.id)
            logger.info(f"Round {round_obj.id} auto-updated to active")
        
        # 2. Update rounds from 'active' to 'completed' when round duration expires
        active_rounds = rounds_query.filter(status='active')
        for round_obj in active_rounds:
            if round_obj.round_start_datetime:
                # Use round_duration_minutes if available, otherwise fallback to duration_minutes
                duration = getattr(round_obj, 'round_duration_minutes', None) or getattr(round_obj, 'duration_minutes', 60)
                end_time = round_obj.round_start_datetime + timezone.timedelta(minutes=duration)
                if now >= end_time:
                    round_obj.status = 'completed'
                    round_obj.save(update_fields=["status", "updated_at"])
                    updated_round_ids.append(round_obj.id)
                    logger.info(f"Round {round_obj.id} auto-updated to completed")
        
        return Response({
            'success': True,
            'updated_round_ids': updated_round_ids,
            'message': f'Updated {len(updated_round_ids)} rounds'
        })
    except Exception as e:
        logger.error(f"Error in auto_update_round_statuses: {str(e)}")
        return Response({
            'success': False,
            'message': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# canadmin/views.py - Updated AddUpdateCodingQuestion

# canadmin/views.py - Fixed AddUpdateCodingQuestion

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def AddUpdateCodingQuestion(request):
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    
    coding_question_id = request.data.get("id")
    round_id = request.data.get("round")
    problem_statement = request.data.get("problem_statement", "").strip()
    
    # Validate round exists
    if not round_id:
        return Response(
            {
                "success": False,
                "message": "Round ID is required."
            },
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        round_obj = RoundModel.objects.get(id=round_id)
    except RoundModel.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "Round not found."
            },
            status=status.HTTP_404_NOT_FOUND
        )
    
    # For new questions, check for duplicates
    if not coding_question_id:
        # Check if a question with the same problem statement already exists in this round
        existing_question = CodingQuestionModel.objects.filter(
            round=round_obj,
            problem_statement__iexact=problem_statement
        ).first()
        
        if existing_question:
            return Response(
                {
                    "success": False,
                    "message": f"A question with the problem statement '{problem_statement}' already exists in this round. Please use a different problem statement."
                },
                status=status.HTTP_400_BAD_REQUEST
            )
    
    # If editing, check for duplicates excluding the current question
    if coding_question_id:
        try:
            coding_question = CodingQuestionModel.objects.get(id=coding_question_id)
        except CodingQuestionModel.DoesNotExist:
            return Response(
                {
                    "success": False,
                    "message": "Coding Question Not Found!!!"
                },
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Check if another question has the same problem statement (excluding this one)
        existing_question = CodingQuestionModel.objects.filter(
            round=round_obj,
            problem_statement__iexact=problem_statement
        ).exclude(id=coding_question_id).first()
        
        if existing_question:
            return Response(
                {
                    "success": False,
                    "message": f"A question with the problem statement '{problem_statement}' already exists in this round. Please use a different problem statement."
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = CodingQuestionSerializer(coding_question, data=request.data, partial=True)
        if serializer.is_valid():
            coding_question = serializer.save()
            return Response(
                {
                    "success": True,
                    "message": "Coding Question Updated Successfully!!!",
                    "data": CodingQuestionSerializer(coding_question).data
                },
                status=status.HTTP_200_OK
            )
        return Response(
            {
                "success": False,
                "message": serializer.errors
            },
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # New question - create
    serializer = CodingQuestionSerializer(data=request.data)
    if serializer.is_valid():
        coding_question = serializer.save()
        return Response(
            {
                "success": True,
                "message": "Coding Question added successfully.",
                "data": CodingQuestionSerializer(coding_question).data
            },
            status=status.HTTP_201_CREATED
        )
    return Response(
        {
            "success": False,
            "message": serializer.errors
        },
        status=status.HTTP_400_BAD_REQUEST
    )

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def GetCodingQuestions(request, round_id):
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    try:
        round_obj = RoundModel.objects.get(id=round_id)
    except RoundModel.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "Round not found."
            },
            status=status.HTTP_404_NOT_FOUND
        )
    if round_obj.round_type != "coding":
        return Response(
            {
                "success": False,
                "message": "This is not an coding round."
            },
            status=status.HTTP_400_BAD_REQUEST
        )
    questions = CodingQuestionModel.objects.filter(round=round_obj)
    serializer = CodingQuestionSerializer(questions, many=True)
    return Response(
        {
            "success": True,
            "message": "Questions fetched successfully.",
            "count": questions.count(),
            "data": serializer.data
        },
        status=status.HTTP_200_OK
    )

@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def DeleteCodingQuestion(request, coding_question_id):
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    try:
        coding_question_obj = CodingQuestionModel.objects.get(id=coding_question_id)
    except CodingQuestionModel.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "Coding Question Not Found."
            },
            status=status.HTTP_404_NOT_FOUND
        )
    coding_question_obj.delete()
    return Response(
        {
            "success": True,
            "message": "Coding Question Deleted Successfully!!",
            "data": {
                "id": coding_question_id
            }
        },
        status=status.HTTP_200_OK
    )

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def AddUpdateCodingTestCase(request):
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    coding_test_case_id = request.data.get("id")
    if coding_test_case_id:
        try:
            coding_test_case = CodingTestCaseModel.objects.get(id = coding_test_case_id)
        except CodingTestCaseModel.DoesNotExist:
            return Response(
                {
                    "success": False,
                    "message": "Coding Test Case Not Found!!!"
                },
                status= status.HTTP_404_NOT_FOUND
            )
        serializer = CodingTestCaseSerializer(coding_test_case, data= request.data, partial=True)
        if serializer.is_valid():
            coding_test_case = serializer.save()
            return Response(
                {
                    "success": True,
                    "message": "Coding Test Case Updated Successfully!!!",
                    "data": CodingTestCaseSerializer(coding_test_case).data
                },
                status= status.HTTP_200_OK
            )
        return Response(
            {
                "success": False,
                "message": serializer.errors
            },
            status= status.HTTP_400_BAD_REQUEST
        )
    serializer = CodingTestCaseSerializer(data=request.data)
    if serializer.is_valid():
        coding_test_case = serializer.save()
        return Response(
            {
                "success": True,
                "message": "Coding Test Case added successfully.",
                "data": CodingTestCaseSerializer(coding_test_case).data
            },
            status=status.HTTP_201_CREATED
        )
    return Response(
        {
            "success": False,
            "message": serializer.errors
        },
        status=status.HTTP_400_BAD_REQUEST
    )

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def GetCodingTestCases(request, coding_question_id):
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    try:
        coding_question_obj = CodingQuestionModel.objects.get(id=coding_question_id)
    except CodingQuestionModel.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "Coding Question not found."
            },
            status=status.HTTP_404_NOT_FOUND
        )
    test_cases = CodingTestCaseModel.objects.filter(question=coding_question_obj)
    serializer = CodingTestCaseSerializer(test_cases, many=True)
    return Response(
        {
            "success": True,
            "message": "Coding Test Cases fetched successfully.",
            "count": test_cases.count(),
            "data": serializer.data
        },
        status=status.HTTP_200_OK
    )

@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def DeleteCodingTestCase(request, coding_test_case_id):
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    try:
        coding_test_case_obj = CodingTestCaseModel.objects.get(id=coding_test_case_id)
    except CodingTestCaseModel.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "Coding Test Case Not Found."
            },
            status=status.HTTP_404_NOT_FOUND
        )
    coding_test_case_obj.delete()
    return Response(
        {
            "success": True,
            "message": "Coding Test Case Deleted Successfully!!",
            "data": {
                "id": coding_test_case_id
            }
        },
        status=status.HTTP_200_OK
    )

# canadmin/views.py - Updated ListCodingResults with detailed question data

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def ListCodingResults(request, round_id):
    """
    Get coding results with question-wise details for a specific attempt
    """
    # --------------------------------------------------
    # 1. ADMIN PERMISSION CHECK
    # --------------------------------------------------
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Only admin can view coding results."
            },
            status=status.HTTP_403_FORBIDDEN
        )

    # --------------------------------------------------
    # 2. GET CODING ROUND
    # --------------------------------------------------
    try:
        round_obj = RoundModel.objects.get(
            id=round_id,
            round_type="coding"
        )
    except RoundModel.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "Coding Round Not Found."
            },
            status=status.HTTP_404_NOT_FOUND
        )

    # Get attempt_id from query params for detailed view
    attempt_id = request.query_params.get('attempt_id')
    
    if attempt_id:
        # --------------------------------------------------
        # Return detailed results for a specific attempt
        # --------------------------------------------------
        try:
            attempt = RoundAttemptModel.objects.select_related(
                'candidate', 'round'
            ).get(
                id=attempt_id,
                round=round_obj
            )
        except RoundAttemptModel.DoesNotExist:
            return Response(
                {
                    "success": False,
                    "message": "Attempt not found."
                },
                status=status.HTTP_404_NOT_FOUND
            )

        # Get coding submission for this attempt
        try:
            coding_submission = CodingSubmissionModel.objects.get(attempt=attempt)
        except CodingSubmissionModel.DoesNotExist:
            coding_submission = None

        # Get question-level submissions
        question_submissions = (
            CodingQuestionSubmissionModel.objects
            .filter(attempt=attempt)
            .select_related('question')
            .order_by('question_id')
        )

        # Get all questions for this round
        questions = CodingQuestionModel.objects.filter(
            round=round_obj
        ).order_by('id')

        # Build question-wise results
        question_results = []
        for question in questions:
            submission = question_submissions.filter(question=question).first()
            
            question_data = {
                'question_id': question.id,
                'question': question.problem_statement,
                'description': question.description,
                'difficulty': question.difficulty,
                'marks': question.marks,
                'language': submission.language if submission else None,
                'code': submission.code if submission else None,
                'status': submission.status if submission else 'not_attempted',
                'score': submission.score if submission else 0,
                'total_test_cases': submission.total_test_cases if submission else 0,
                'passed_test_cases': submission.passed_test_cases if submission else 0,
                'submitted_at': submission.submitted_at if submission else None,
                'evaluated_at': submission.evaluated_at if submission else None,
            }
            question_results.append(question_data)

        # Get candidate decision
        candidate_decision = None
        try:
            decision_obj = RoundCandidateDecisionModel.objects.get(attempt=attempt)
            candidate_decision = {
                'decision': decision_obj.decision,
                'score': decision_obj.score,
                'total_marks': decision_obj.total_marks,
                'percentage': float(decision_obj.percentage) if decision_obj.percentage else 0
            }
        except RoundCandidateDecisionModel.DoesNotExist:
            pass

        # Calculate stats
        total_questions = len(question_results)
        attempted_questions = sum(1 for q in question_results if q['status'] != 'not_attempted')
        submitted_questions = sum(1 for q in question_results if q['status'] == 'submitted' or q['status'] == 'evaluated')
        total_score = sum(q['score'] for q in question_results)
        total_marks = sum(q['marks'] for q in question_results)
        
        # Calculate test case stats
        total_test_cases = sum(q['total_test_cases'] for q in question_results)
        passed_test_cases = sum(q['passed_test_cases'] for q in question_results)
        
        # Get coding submission data
        submission_data = None
        if coding_submission:
            submission_data = {
                'submission_id': coding_submission.id,
                'status': coding_submission.status,
                'total_questions': coding_submission.total_questions,
                'attempted_questions': coding_submission.attempted_questions,
                'total_marks': coding_submission.total_marks,
                'score': coding_submission.score,
                'submitted_at': coding_submission.submitted_at,
                'evaluated_at': coding_submission.evaluated_at
            }

        return Response(
            {
                "success": True,
                "message": "Coding results fetched successfully.",
                "data": {
                    "attempt_id": attempt.id,
                    "candidate_name": attempt.candidate.name or attempt.candidate.email,
                    "candidate_email": attempt.candidate.email,
                    "score": total_score,
                    "total_marks": total_marks,
                    "percentage": round((total_score / total_marks * 100), 2) if total_marks > 0 else 0,
                    "total_questions": total_questions,
                    "attempted_questions": attempted_questions,
                    "submitted_questions": submitted_questions,
                    "total_test_cases": total_test_cases,
                    "passed_test_cases": passed_test_cases,
                    "status": attempt.status,
                    "submitted_at": attempt.submitted_at,
                    "candidate_decision": candidate_decision,
                    "coding_submission": submission_data,
                    "questions": question_results
                }
            },
            status=status.HTTP_200_OK
        )

    # --------------------------------------------------
    # 3. GET FINAL CODING SUBMISSIONS (List View)
    # --------------------------------------------------
    submissions = (
        CodingSubmissionModel.objects
        .filter(attempt__round=round_obj)
        .select_related(
            "attempt",
            "attempt__candidate"
        )
        .prefetch_related(
            Prefetch(
                "attempt__coding_question_submissions",
                queryset=CodingQuestionSubmissionModel.objects
                .select_related("question")
                .order_by("question_id")
            )
        )
        .order_by("-submitted_at")
    )

    # --------------------------------------------------
    # 4. IF NO FINAL SUBMISSIONS
    # --------------------------------------------------
    if not submissions.exists():
        return Response(
            {
                "success": True,
                "message": "No coding submissions found for this round.",
                "data": {
                    "round_id": round_obj.id,
                    "round_type": round_obj.round_type,
                    "round_order": round_obj.round_order,
                    "total_candidates": 0,
                    "results": []
                }
            },
            status=status.HTTP_200_OK
        )

    # --------------------------------------------------
    # 5. GET ALL DECISIONS IN ONE QUERY
    # --------------------------------------------------
    attempt_ids = [submission.attempt.id for submission in submissions]
    decisions = RoundCandidateDecisionModel.objects.filter(
        attempt_id__in=attempt_ids
    )
    decision_map = {decision.attempt_id: decision for decision in decisions}

    # --------------------------------------------------
    # 6. PREPARE RESULTS
    # --------------------------------------------------
    results = []

    for submission in submissions:
        attempt = submission.attempt
        candidate = attempt.candidate

        # ----------------------------------------------
        # Final round score
        # ----------------------------------------------
        score = submission.score or 0
        total_marks = submission.total_marks or 0

        percentage = 0
        if total_marks > 0:
            percentage = round(
                (score / total_marks) * 100,
                2
            )

        # ----------------------------------------------
        # Candidate information
        # ----------------------------------------------
        candidate_name = getattr(
            candidate,
            "name",
            None
        )
        if not candidate_name:
            candidate_name = candidate.email

        # ----------------------------------------------
        # Get Candidate Decision from the map
        # ----------------------------------------------
        candidate_decision = None
        decision_obj = decision_map.get(attempt.id)
        if decision_obj:
            candidate_decision = {
                "decision": decision_obj.decision,
                "score": decision_obj.score,
                "total_marks": decision_obj.total_marks,
                "percentage": float(decision_obj.percentage) if decision_obj.percentage else 0
            }

        # ----------------------------------------------
        # Question-level submissions (summary for list view)
        # ----------------------------------------------
        question_results = []
        question_submissions = (
            attempt.coding_question_submissions.all()
        )

        for question_submission in question_submissions:
            question_score = question_submission.score or 0

            question_results.append(
                {
                    "submission_id": question_submission.id,
                    "question_id": question_submission.question.id,
                    "question": (
                        question_submission.question.problem_statement
                        or "N/A"
                    ),
                    "language": question_submission.language,
                    "score": question_score,
                    "total_test_cases": (
                        question_submission.total_test_cases or 0
                    ),
                    "passed_test_cases": (
                        question_submission.passed_test_cases or 0
                    ),
                    "status": question_submission.status,
                    "submitted_at": question_submission.submitted_at,
                    "evaluated_at": question_submission.evaluated_at,
                }
            )

        # ----------------------------------------------
        # Final candidate result
        # ----------------------------------------------
        results.append(
            {
                "submission_id": submission.id,
                "attempt_id": attempt.id,
                "candidate_id": candidate.id,
                "candidate_name": candidate_name,
                "candidate_email": candidate.email,
                "status": submission.status,
                "score": score,
                "total_marks": total_marks,
                "percentage": percentage,
                "total_questions": (
                    submission.total_questions or 0
                ),
                "attempted_questions": (
                    submission.attempted_questions or 0
                ),
                "submitted_at": submission.submitted_at,
                "evaluated_at": submission.evaluated_at,
                "questions": question_results,
                "candidate_decision": candidate_decision,
            }
        )

    # --------------------------------------------------
    # 7. CALCULATE SUMMARY
    # --------------------------------------------------
    total_candidates = len(results)
    passed_count = 0
    failed_count = 0
    pending_count = 0

    for result in results:
        decision = result.get("candidate_decision")
        if decision:
            if decision["decision"] == "shortlisted":
                passed_count += 1
            elif decision["decision"] == "rejected":
                failed_count += 1
            else:
                pending_count += 1
        else:
            pending_count += 1

    # --------------------------------------------------
    # 8. FINAL RESPONSE
    # --------------------------------------------------
    return Response(
        {
            "success": True,
            "message": "Coding Results Fetched Successfully.",
            "data": {
                "round_id": round_obj.id,
                "round_type": round_obj.round_type,
                "round_order": round_obj.round_order,
                "total_candidates": total_candidates,
                "results": results,
                "summary": {
                    "passed": passed_count,
                    "failed": failed_count,
                    "pending": pending_count
                }
            }
        },
        status=status.HTTP_200_OK
    )

# canadmin/views.py - Updated GetRoundDetails

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def GetRoundDetails(request, round_id):
    """
    Get detailed information about a specific round.

    Aptitude:
        - Questions/marks from AptitudeQuestionModel
        - Candidate results from RoundAttemptModel

    Coding:
        - Questions/marks from CodingQuestionModel
        - Candidate attempts from RoundAttemptModel
        - Final candidate results from CodingSubmissionModel
        - Question-level results from CodingQuestionSubmissionModel
    """

    # ---------------------------------------------------------
    # 1. ADMIN PERMISSION CHECK
    # ---------------------------------------------------------
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied. Only admin can view round details."
            },
            status=status.HTTP_403_FORBIDDEN
        )

    try:
        # ---------------------------------------------------------
        # 2. GET ROUND
        # ---------------------------------------------------------
        round_obj = (
            RoundModel.objects
            .select_related("drive")
            .get(id=round_id)
        )

        # ---------------------------------------------------------
        # 3. GET QUESTION STATISTICS
        # ---------------------------------------------------------
        total_questions = 0
        total_marks = 0

        if round_obj.round_type == "aptitude":

            questions = AptitudeQuestionModel.objects.filter(
                round=round_obj
            )

            total_questions = questions.count()

            total_marks = (
                questions.aggregate(
                    total=Sum("marks")
                )["total"] or 0
            )

        elif round_obj.round_type == "coding":

            questions = CodingQuestionModel.objects.filter(
                round=round_obj
            )

            total_questions = questions.count()

            total_marks = (
                questions.aggregate(
                    total=Sum("marks")
                )["total"] or 0
            )

        # ---------------------------------------------------------
        # 4. GET ALL CANDIDATE ATTEMPTS
        # ---------------------------------------------------------
        attempts = (
            RoundAttemptModel.objects
            .filter(round=round_obj)
            .select_related("candidate")
        )

        total_candidates = attempts.count()

        # ---------------------------------------------------------
        # 5. COMMON STATISTICS
        # ---------------------------------------------------------
        completed = 0
        pending = 0
        passed = 0
        failed = 0

        scores = []

        passing_percentage = getattr(
            round_obj,
            "passing_percentage",
            40
        )

        # Get round duration (handle both field names)
        round_duration = getattr(round_obj, 'round_duration_minutes', None) or getattr(round_obj, 'duration_minutes', 60)
        test_duration = getattr(round_obj, 'test_duration_minutes', None) or round_duration

        # =========================================================
        # 6. APTITUDE ROUND
        # =========================================================
        if round_obj.round_type == "aptitude":

            for attempt in attempts:

                score = attempt.score or 0

                attempt_total_marks = (
                    attempt.total_marks
                    or total_marks
                )

                # -------------------------------------------------
                # Pending / In Progress
                # -------------------------------------------------
                if attempt.status == "in_progress":
                    pending += 1
                    continue

                # -------------------------------------------------
                # Completed / Evaluated / Passed / Failed
                # -------------------------------------------------
                if attempt.status in [
                    "completed",
                    "evaluated",
                    "passed",
                    "failed"
                ]:
                    completed += 1

                    scores.append(score)

                    # Calculate percentage
                    percentage = 0

                    if attempt_total_marks > 0:
                        percentage = (
                            score / attempt_total_marks
                        ) * 100

                    # -------------------------------------------------
                    # Passed / Failed
                    # -------------------------------------------------
                    if attempt.status == "passed":
                        passed += 1

                    elif attempt.status == "failed":
                        failed += 1

                    else:
                        if percentage >= passing_percentage:
                            passed += 1
                        else:
                            failed += 1

        # =========================================================
        # 7. CODING ROUND
        # =========================================================
        elif round_obj.round_type == "coding":

            # -----------------------------------------------------
            # Get final coding submissions
            #
            # CodingSubmissionModel represents the FINAL
            # submission of the coding round.
            # -----------------------------------------------------
            coding_submissions = (
                CodingSubmissionModel.objects
                .filter(
                    attempt__round=round_obj
                )
                .select_related("attempt")
            )

            # -----------------------------------------------------
            # IDs of attempts that have submitted the round
            # -----------------------------------------------------
            submitted_attempt_ids = set(
                coding_submissions.values_list(
                    "attempt_id",
                    flat=True
                )
            )

            # -----------------------------------------------------
            # Pending candidates
            #
            # Candidate has an attempt but has not submitted
            # the complete coding round yet.
            # -----------------------------------------------------
            pending = attempts.filter(
                status="in_progress"
            ).exclude(
                id__in=submitted_attempt_ids
            ).count()

            # -----------------------------------------------------
            # Process final coding submissions
            # -----------------------------------------------------
            for submission in coding_submissions:

                score = submission.score or 0

                submission_total_marks = (
                    submission.total_marks
                    or total_marks
                )

                completed += 1

                scores.append(score)

                # -------------------------------------------------
                # Calculate percentage
                # -------------------------------------------------
                percentage = 0

                if submission_total_marks > 0:
                    percentage = (
                        score / submission_total_marks
                    ) * 100

                # -------------------------------------------------
                # Passed / Failed
                # -------------------------------------------------
                if submission.status == "passed":
                    passed += 1

                elif submission.status == "failed":
                    failed += 1

                else:
                    # Fallback if final submission status has not
                    # yet been updated to passed/failed.
                    if percentage >= passing_percentage:
                        passed += 1
                    else:
                        failed += 1

        # =========================================================
        # 8. SCORE STATISTICS
        # =========================================================
        average_score = (
            round(
                sum(scores) / len(scores),
                1
            )
            if scores
            else 0
        )

        highest_score = (
            max(scores)
            if scores
            else 0
        )

        lowest_score = (
            min(scores)
            if scores
            else 0
        )

        # =========================================================
        # 9. RESPONSE DATA
        # =========================================================
        data = {
            "id": round_obj.id,

            "round_type": round_obj.round_type,

            "round_type_display": (
                round_obj.get_round_type_display()
            ),

            "round_order": round_obj.round_order,

            "status": round_obj.status,

            "round_duration_minutes": round_duration,
            "test_duration_minutes": test_duration,
            "duration_minutes": round_duration,  # For backward compatibility

            "passing_percentage": passing_percentage,

            "created_at": round_obj.created_at,

            "total_questions": total_questions,

            "total_marks": total_marks,

            "drive_id": (
                round_obj.drive.id
                if round_obj.drive
                else None
            ),

            "drive_title": (
                round_obj.drive.title
                if round_obj.drive
                else None
            ),

            "candidate_stats": {

                "total_candidates": total_candidates,

                "completed": completed,

                "pending": pending,

                "passed": passed,

                "failed": failed,

                "average_score": average_score,

                "highest_score": highest_score,

                "lowest_score": lowest_score
            }
        }

        # =========================================================
        # 10. SUCCESS RESPONSE
        # =========================================================
        return Response(
            {
                "success": True,
                "message": "Round details fetched successfully.",
                "data": data
            },
            status=status.HTTP_200_OK
        )

    except RoundModel.DoesNotExist:

        return Response(
            {
                "success": False,
                "message": "Round not found."
            },
            status=status.HTTP_404_NOT_FOUND
        )

    except Exception as e:

        logger.error(
            f"Error in GetRoundDetails: {str(e)}",
            exc_info=True
        )

        return Response(
            {
                "success": False,
                "message": f"An error occurred: {str(e)}"
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def PreviewRoundResults(request, round_id):
    """
    Preview passed/failed results based on an admin-provided
    passing percentage.

    IMPORTANT:
    This API DOES NOT update the database.
    It only returns the suggested results.
    """

    # ---------------------------------------------------------
    # 1. ADMIN PERMISSION
    # ---------------------------------------------------------
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied. Only admin can mark results."
            },
            status=status.HTTP_403_FORBIDDEN
        )

    # ---------------------------------------------------------
    # 2. GET ROUND
    # ---------------------------------------------------------
    try:
        round_obj = RoundModel.objects.get(id=round_id)
    except RoundModel.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "Round not found."
            },
            status=status.HTTP_404_NOT_FOUND
        )

    # ---------------------------------------------------------
    # 3. ONLY APTITUDE / CODING
    # ---------------------------------------------------------
    if round_obj.round_type not in ["aptitude", "coding"]:
        return Response(
            {
                "success": False,
                "message": (
                    "Result marking is currently supported only "
                    "for aptitude and coding rounds."
                )
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # ---------------------------------------------------------
    # 4. GET PASSING PERCENTAGE
    # ---------------------------------------------------------
    passing_percentage = request.data.get("passing_percentage")

    if passing_percentage is None:
        return Response(
            {
                "success": False,
                "message": "Passing percentage is required."
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        passing_percentage = float(passing_percentage)
    except (TypeError, ValueError):
        return Response(
            {
                "success": False,
                "message": "Passing percentage must be a valid number."
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # ---------------------------------------------------------
    # 5. VALIDATE RANGE
    # ---------------------------------------------------------
    if passing_percentage < 0 or passing_percentage > 100:
        return Response(
            {
                "success": False,
                "message": "Passing percentage must be between 0 and 100."
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    results = []

    # =========================================================
    # 6. APTITUDE
    # =========================================================
    if round_obj.round_type == "aptitude":

        attempts = (
            RoundAttemptModel.objects
            .filter(round=round_obj)
            .select_related("candidate")
            .order_by("-submitted_at", "-started_at")
        )

        # ---------------------------------------------------------
        # Get existing decisions for all attempts
        # ---------------------------------------------------------
        attempt_ids = [attempt.id for attempt in attempts]
        decisions = RoundCandidateDecisionModel.objects.filter(
            attempt_id__in=attempt_ids
        )
        decision_map = {decision.attempt_id: decision for decision in decisions}

        for attempt in attempts:

            score = attempt.score or 0
            total_marks = attempt.total_marks or 0

            percentage = 0

            if total_marks > 0:
                percentage = round(
                    (score / total_marks) * 100,
                    2
                )

            # Only completed/evaluated attempts
            if attempt.status == "in_progress":
                suggested_result = "pending"
            else:
                suggested_result = (
                    "passed"
                    if percentage >= passing_percentage
                    else "failed"
                )

            candidate_name = getattr(
                attempt.candidate,
                "name",
                None
            )

            if not candidate_name:
                candidate_name = attempt.candidate.email

            # ---------------------------------------------------------
            # Get existing candidate decision
            # ---------------------------------------------------------
            candidate_decision = None
            decision_obj = decision_map.get(attempt.id)
            if decision_obj:
                candidate_decision = {
                    "decision": decision_obj.decision,
                    "score": decision_obj.score,
                    "total_marks": decision_obj.total_marks,
                    "percentage": float(decision_obj.percentage) if decision_obj.percentage else 0
                }

            results.append(
                {
                    "attempt_id": attempt.id,
                    "candidate_id": attempt.candidate.id,
                    "candidate_name": candidate_name,
                    "candidate_email": attempt.candidate.email,

                    "score": score,
                    "total_marks": total_marks,
                    "percentage": percentage,

                    "current_status": attempt.status,
                    "suggested_result": suggested_result,
                    "candidate_decision": candidate_decision,  # <-- ADD THIS
                }
            )

    # =========================================================
    # 7. CODING
    # =========================================================
    elif round_obj.round_type == "coding":

        submissions = (
            CodingSubmissionModel.objects
            .filter(attempt__round=round_obj)
            .select_related(
                "attempt",
                "attempt__candidate"
            )
            .order_by("-submitted_at")
        )

        # ---------------------------------------------------------
        # Get existing decisions for all attempts
        # ---------------------------------------------------------
        attempt_ids = [submission.attempt.id for submission in submissions]
        decisions = RoundCandidateDecisionModel.objects.filter(
            attempt_id__in=attempt_ids
        )
        decision_map = {decision.attempt_id: decision for decision in decisions}

        for submission in submissions:

            candidate = submission.attempt.candidate

            score = submission.score or 0
            total_marks = submission.total_marks or 0

            percentage = 0

            if total_marks > 0:
                percentage = round(
                    (score / total_marks) * 100,
                    2
                )

            suggested_result = (
                "passed"
                if percentage >= passing_percentage
                else "failed"
            )

            candidate_name = getattr(
                candidate,
                "name",
                None
            )

            if not candidate_name:
                candidate_name = candidate.email

            # ---------------------------------------------------------
            # Get existing candidate decision
            # ---------------------------------------------------------
            candidate_decision = None
            decision_obj = decision_map.get(submission.attempt.id)
            if decision_obj:
                candidate_decision = {
                    "decision": decision_obj.decision,
                    "score": decision_obj.score,
                    "total_marks": decision_obj.total_marks,
                    "percentage": float(decision_obj.percentage) if decision_obj.percentage else 0
                }

            results.append(
                {
                    "submission_id": submission.id,
                    "attempt_id": submission.attempt.id,

                    "candidate_id": candidate.id,
                    "candidate_name": candidate_name,
                    "candidate_email": candidate.email,

                    "score": score,
                    "total_marks": total_marks,
                    "percentage": percentage,

                    "current_status": submission.status,
                    "suggested_result": suggested_result,
                    "candidate_decision": candidate_decision,  # <-- ADD THIS
                }
            )

    # ---------------------------------------------------------
    # 8. SUMMARY
    # ---------------------------------------------------------
    passed_count = len([
        result
        for result in results
        if result["suggested_result"] == "passed"
    ])

    failed_count = len([
        result
        for result in results
        if result["suggested_result"] == "failed"
    ])

    pending_count = len([
        result
        for result in results
        if result["suggested_result"] == "pending"
    ])

    # ---------------------------------------------------------
    # 9. RESPONSE
    # ---------------------------------------------------------
    return Response(
        {
            "success": True,
            "message": "Result preview generated successfully.",

            "data": {
                "round_id": round_obj.id,
                "round_type": round_obj.round_type,
                "passing_percentage": passing_percentage,

                "summary": {
                    "total_candidates": len(results),
                    "suggested_passed": passed_count,
                    "suggested_failed": failed_count,
                    "pending": pending_count
                },

                "results": results
            }
        },
        status=status.HTTP_200_OK
    )

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def ConfirmRoundResults(request, round_id):
    """
    Confirm and save final passed/failed results.

    Admin can manually override the previewed result
    before sending this request.

    Supports both Aptitude and Coding rounds.
    """

    # ---------------------------------------------------------
    # 1. ADMIN PERMISSION
    # ---------------------------------------------------------
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied. Only admin can mark results."
            },
            status=status.HTTP_403_FORBIDDEN
        )

    # ---------------------------------------------------------
    # 2. GET ROUND
    # ---------------------------------------------------------
    try:
        round_obj = RoundModel.objects.get(id=round_id)
    except RoundModel.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "Round not found."
            },
            status=status.HTTP_404_NOT_FOUND
        )

    # ---------------------------------------------------------
    # 3. VALIDATE ROUND TYPE
    # ---------------------------------------------------------
    if round_obj.round_type not in ["aptitude", "coding"]:
        return Response(
            {
                "success": False,
                "message": (
                    "Result marking is currently supported only "
                    "for aptitude and coding rounds."
                )
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # ---------------------------------------------------------
    # 4. GET RESULTS
    # ---------------------------------------------------------
    results_data = request.data.get("results")

    if not isinstance(results_data, list) or not results_data:
        return Response(
            {
                "success": False,
                "message": "Results must be a non-empty list."
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    # ---------------------------------------------------------
    # 5. VALID RESULT VALUES
    # ---------------------------------------------------------
    allowed_results = {"passed", "failed"}

    # ---------------------------------------------------------
    # 6. UPDATE INSIDE TRANSACTION
    # ---------------------------------------------------------
    updated_results = []

    try:

        with transaction.atomic():

            # =================================================
            # APTITUDE ROUND
            # =================================================
            if round_obj.round_type == "aptitude":

                for item in results_data:

                    attempt_id = item.get("attempt_id")
                    result = item.get("result")

                    if not attempt_id:
                        raise ValueError("attempt_id is required for aptitude result.")

                    if result not in allowed_results:
                        raise ValueError(f"Result must be either 'passed' or 'failed'. Got: {result}")

                    # Get the attempt
                    try:
                        attempt = RoundAttemptModel.objects.select_related("candidate").get(
                            id=attempt_id,
                            round=round_obj
                        )
                    except RoundAttemptModel.DoesNotExist:
                        raise ValueError(f"Attempt {attempt_id} does not belong to this round.")

                    # -----------------------------------------
                    # Update attempt status
                    # -----------------------------------------
                    attempt.status = result
                    attempt.save(update_fields=["status"])

                    # -----------------------------------------
                    # Calculate values for RoundCandidateDecisionModel
                    # -----------------------------------------
                    score = attempt.score or 0
                    total_marks = attempt.total_marks or 0
                    percentage = 0
                    if total_marks > 0:
                        percentage = round((score / total_marks) * 100, 2)

                    # Map frontend result to decision
                    decision = "shortlisted" if result == "passed" else "rejected"

                    # -----------------------------------------
                    # CREATE/UPDATE RoundCandidateDecisionModel
                    # -----------------------------------------
                    decision_obj, created = RoundCandidateDecisionModel.objects.update_or_create(
                        attempt=attempt,
                        defaults={
                            'decision': decision,
                            'score': score,
                            'total_marks': total_marks,
                            'percentage': percentage
                        }
                    )

                    # Get candidate name
                    candidate_name = getattr(attempt.candidate, "name", None)
                    if not candidate_name:
                        candidate_name = attempt.candidate.email

                    updated_results.append({
                        "attempt_id": attempt.id,
                        "candidate_id": attempt.candidate.id,
                        "candidate_name": candidate_name,
                        "status": attempt.status,
                        "decision": decision_obj.decision,
                        "percentage": float(decision_obj.percentage),
                        "is_evaluated": True
                    })

            # =================================================
            # CODING ROUND
            # =================================================
            elif round_obj.round_type == "coding":

                for item in results_data:

                    submission_id = item.get("submission_id")
                    result = item.get("result")

                    if not submission_id:
                        raise ValueError("submission_id is required for coding result.")

                    if result not in allowed_results:
                        raise ValueError(f"Result must be either 'passed' or 'failed'. Got: {result}")

                    # Get the submission
                    try:
                        submission = CodingSubmissionModel.objects.select_related(
                            "attempt",
                            "attempt__candidate"
                        ).get(
                            id=submission_id,
                            attempt__round=round_obj
                        )
                    except CodingSubmissionModel.DoesNotExist:
                        raise ValueError(f"Submission {submission_id} does not belong to this round.")

                    # -----------------------------------------
                    # Update submission status
                    # -----------------------------------------
                    submission.status = result
                    submission.save(update_fields=["status"])

                    # -----------------------------------------
                    # Calculate values for RoundCandidateDecisionModel
                    # -----------------------------------------
                    attempt = submission.attempt
                    score = submission.score or 0
                    total_marks = submission.total_marks or 0
                    percentage = 0
                    if total_marks > 0:
                        percentage = round((score / total_marks) * 100, 2)

                    # Map frontend result to decision
                    decision = "shortlisted" if result == "passed" else "rejected"

                    # -----------------------------------------
                    # CREATE/UPDATE RoundCandidateDecisionModel
                    # -----------------------------------------
                    decision_obj, created = RoundCandidateDecisionModel.objects.update_or_create(
                        attempt=attempt,
                        defaults={
                            'decision': decision,
                            'score': score,
                            'total_marks': total_marks,
                            'percentage': percentage
                        }
                    )

                    # Get candidate name
                    candidate = attempt.candidate
                    candidate_name = getattr(candidate, "name", None)
                    if not candidate_name:
                        candidate_name = candidate.email

                    updated_results.append({
                        "submission_id": submission.id,
                        "attempt_id": attempt.id,
                        "candidate_id": candidate.id,
                        "candidate_name": candidate_name,
                        "status": submission.status,
                        "decision": decision_obj.decision,
                        "percentage": float(decision_obj.percentage),
                        "is_evaluated": True
                    })

    except ValueError as e:
        return Response(
            {
                "success": False,
                "message": str(e)
            },
            status=status.HTTP_400_BAD_REQUEST
        )

    except Exception as e:
        logger.error(f"Error confirming round results: {str(e)}", exc_info=True)
        return Response(
            {
                "success": False,
                "message": f"An error occurred: {str(e)}"
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

    # ---------------------------------------------------------
    # 7. RESPONSE
    # ---------------------------------------------------------
    return Response(
        {
            "success": True,
            "message": "Round results marked successfully.",
            "data": {
                "round_id": round_obj.id,
                "round_type": round_obj.round_type,
                "updated_count": len(updated_results),
                "results": updated_results
            }
        },
        status=status.HTTP_200_OK
    )

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def GetUserDriveAttempts(request, user_id):
    """
    Get all drive attempts and round results for a specific user
    """
    if request.user.role != "admin":
        return Response(
            {
                "success": False,
                "message": "Permission denied."
            },
            status=status.HTTP_403_FORBIDDEN
        )
    
    try:
        user = UserTable.objects.get(id=user_id, role="candidate")
    except UserTable.DoesNotExist:
        return Response(
            {
                "success": False,
                "message": "User Not Found"
            },
            status=status.HTTP_404_NOT_FOUND
        )
    
    # Get all attempts for this user
    attempts = RoundAttemptModel.objects.filter(
        candidate=user
    ).select_related(
        'round',
        'round__drive',
        'round__drive__institute'
    ).order_by('-round__drive__drive_date_time', 'round__round_order')
    
    # Group attempts by drive
    drives_data = {}
    for attempt in attempts:
        drive = attempt.round.drive
        drive_id = drive.id
        
        if drive_id not in drives_data:
            drives_data[drive_id] = {
                'drive_id': drive.id,
                'drive_title': drive.title,
                'drive_status': drive.status,
                'drive_date_time': drive.drive_date_time,
                'job_role': drive.job_role,
                'ctc': drive.ctc,
                'job_location': drive.job_location,
                'rounds': []
            }
        
        # Get candidate decision if exists
        decision = None
        try:
            decision_obj = RoundCandidateDecisionModel.objects.get(attempt=attempt)
            decision = {
                'decision': decision_obj.decision,
                'score': decision_obj.score,
                'total_marks': decision_obj.total_marks,
                'percentage': float(decision_obj.percentage) if decision_obj.percentage else 0
            }
        except RoundCandidateDecisionModel.DoesNotExist:
            pass
        
        # Get coding submission if applicable
        coding_submission = None
        if attempt.round.round_type == 'coding':
            try:
                coding_sub = CodingSubmissionModel.objects.get(attempt=attempt)
                coding_submission = {
                    'status': coding_sub.status,
                    'score': coding_sub.score,
                    'total_marks': coding_sub.total_marks,
                    'total_questions': coding_sub.total_questions,
                    'attempted_questions': coding_sub.attempted_questions,
                    'submitted_at': coding_sub.submitted_at
                }
            except CodingSubmissionModel.DoesNotExist:
                pass
        
        round_data = {
            'round_id': attempt.round.id,
            'attempt_id': attempt.id,  # <-- THIS IS THE KEY CHANGE
            'round_type': attempt.round.round_type,
            'round_type_display': attempt.round.get_round_type_display(),
            'round_order': attempt.round.round_order,
            'round_status': attempt.round.status,
            'attempt_status': attempt.status,
            'score': attempt.score,
            'total_marks': attempt.total_marks,
            'percentage': round((attempt.score / attempt.total_marks * 100), 2) if attempt.total_marks > 0 else 0,
            'started_at': attempt.started_at,
            'submitted_at': attempt.submitted_at,
            'decision': decision,
            'coding_submission': coding_submission
        }
        
        drives_data[drive_id]['rounds'].append(round_data)
    
    # Convert to list and sort by drive date
    result = list(drives_data.values())
    result.sort(key=lambda x: x['drive_date_time'], reverse=True)
    
    return Response(
        {
            "success": True,
            "message": "User drive attempts fetched successfully!",
            "data": result
        },
        status=status.HTTP_200_OK
    )